import os
from datetime import datetime
from openpyxl import load_workbook
from config import TRACKER_PATH

COLUMNS = [
    "#", "Position Name", "Position Type", "Submission Date", "Date Found",
    "Status", "Vendor Name", "Recruiter", "Email", "Phone Number",
    "Company / End Client", "Location", "Work Model", "Resume Used",
    "Link", "Job Folder", "Interview Date", "Salary / Rate",
    "Next Steps", "Notes"
]

def get_next_number(ws):
    row = ws.max_row
    while row > 2:
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip() != "":
            try:
                return int(val) + 1
            except:
                pass
        row -= 1
    return 1

def log_application(job_title, company, platform, url, resume_filename, job_type, location, work_model, salary=""):
    print(f"  📊 Logging application to tracker...")

    if not os.path.exists(TRACKER_PATH):
        print(f"  ⚠️  Tracker not found at {TRACKER_PATH}")
        return

    wb = load_workbook(TRACKER_PATH)
    ws = wb["Job Applications"]

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    next_num = get_next_number(ws)
    job_folder = f"Desktop/Job Applications/{company.replace(' ', '_')}_{platform.capitalize()}"

    new_row = {
        "#": next_num,
        "Position Name": job_title,
        "Position Type": job_type,
        "Submission Date": date_str,
        "Date Found": date_str,
        "Status": "Applied",
        "Vendor Name": "",
        "Recruiter": "",
        "Email": "",
        "Phone Number": "",
        "Company / End Client": company,
        "Location": location,
        "Work Model": work_model,
        "Resume Used": resume_filename,
        "Link": url,
        "Job Folder": job_folder,
        "Interview Date": "",
        "Salary / Rate": salary,
        "Next Steps": "Follow up in 1 week",
        "Notes": f"Auto-applied via bot — {platform.capitalize()}"
    }

    header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "#":
                header_row = cell.row
                break
        if header_row:
            break

    if not header_row:
        print("  ⚠️  Could not find header row in tracker")
        return

    col_map = {}
    for cell in ws[header_row]:
        if cell.value in COLUMNS:
            col_map[cell.value] = cell.column

    # Find the actual last row with data (max_row includes pre-formatted empty rows)
    next_data_row = ws.max_row
    while next_data_row > header_row:
        row_vals = [ws.cell(row=next_data_row, column=c).value for c in col_map.values()]
        if any(v is not None and str(v).strip() != "" for v in row_vals):
            next_data_row += 1
            break
        next_data_row -= 1
    else:
        next_data_row = header_row + 1

    for col_name, value in new_row.items():
        if col_name in col_map:
            ws.cell(row=next_data_row, column=col_map[col_name], value=value)

    wb.save(TRACKER_PATH)
    print(f"  ✅ Logged: {job_title} at {company} (row {next_data_row})")